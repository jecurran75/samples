from jctools import mytools
import openpyxl


########################################################################

sub_dir = "c:\\pystuff\\sjc10\\"

########################################################################

def make_arp_xlsx(filename):
    wb = openpyxl.Workbook()
    wb.create_sheet(index=1, title="arp_list")
    sheet = wb.get_sheet_by_name("arp_list")
    for i,line in enumerate(inscope_matrix):
        cell = "A"+str(i+1)
        sheet[cell] = line.address
        cell = "B"+str(i+1)
        sheet[cell] = line.MAC
        cell = "C"+str(i+1)
        sheet[cell] = line.switch
        if hasattr(line,"switchport"):
            cell = "D"+str(i+1)
            sheet[cell] = line.switchport
            
    wb.create_sheet(index=2, title="out_of_scope")
    sheet = wb.get_sheet_by_name("out_of_scope")   
    for i,line in enumerate(out_of_scope_matrix):
        cell = "A"+str(i+1)
        sheet[cell] = line.address
        cell = "B"+str(i+1)
        sheet[cell] = line.MAC             
            
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(filename)
    
    ########################################################################

def make_arp_list(ARP_List):
    ObjList = []
    for item in ARP_List:
        IP = item.split()[1]
        MAC = item.split()[3]
        object = mytools.IPv4(IP)
        object.MAC = MAC
        if object.MAC.count(".") == 2 and len(object.MAC) == 14:
            ObjList.append(object)
    return ObjList
    
########################################################################

def get_arp_list():
    r1 = mytools.NetDev("sjc10-dt-gw1")
    r1.login(quiet=True)
    arp_list = r1.sendcmd("sh ip arp",sleep=3,buff=500000)[2:-1]
    r1.logout(quiet=True)
    print "Arp Table taken from device: %s" % r1.name
    return arp_list
    
##############################################################################

def arp_matrix():
    arp_table = get_arp_list()
    arp_matrix = make_arp_list(arp_table)
    
    print
    print "%d ARP entires found." % len(arp_matrix)

    return arp_matrix
    
      
##############################################################################  
    
arp_matrix = arp_matrix()

    
switches = mytools.read_txt(sub_dir+"sjc10-inputs\\in-scope.txt")

for device in switches:
    switch = mytools.NetDev(device)
    switch.login()
    mac_table = switch.sendcmd("sh mac add",buff=1500000)
    switch.logout()
    
    for item in arp_matrix:
        for line in mac_table:
            if item.MAC in line and line.split()[-1] != "Port-channel1":
                item.switch = device
                item.switchport = line.split()[-1]

dt = mytools.NetDev("sjc10-dt-gw1")
dt.login()
wlc1 = dt.sendcmd("sho mac add int g1/1/17",buff=1500000)
wlc2 = dt.sendcmd("sho mac add int g2/1/17",buff=1500000)
wlc3 = dt.sendcmd("sho mac add int po103",buff=1500000)
wlc4 = dt.sendcmd("sho mac add int po104",buff=1500000)     
dt.logout()

for item in arp_matrix:
    for line in wlc1:
        if item.MAC in line:
            item.switch = "sjc10-wl-wlc1"    
    for line in wlc2:
        if item.MAC in line:
            item.switch = "sjc10-wl-wlc2"
    for line in wlc3:
        if item.MAC in line:
            item.switch = "sjc10-wl-wlc3"
    for line in wlc4:
        if item.MAC in line:
            item.switch = "sjc10-wl-wlc4"
        
                
inscope_matrix = [item for item in arp_matrix if hasattr(item,"switch")]
out_of_scope_matrix = [item for item in arp_matrix if not hasattr(item,"switch")]
         
make_arp_xlsx(sub_dir+"sjc10_arplist.xlsx")



