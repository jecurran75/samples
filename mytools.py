#Main tools module
#By James Curran

import openpyxl
import socket
from subprocess import call, PIPE, STDOUT, Popen
import paramiko
import time
import telnetlib
import cPickle as pickle
from getpass import getpass,getuser

#######################################################

#This function pings an IP to see if it's alive.  If dc (double check) is true,
#Try a second time if the first ping fails.

def is_network_alive(item,dc=False):
    args = ("ping -n 1 "+item).split()    
    if call(args, stdout=PIPE, stderr=STDOUT, creationflags=8) == 0:
        return True
    elif dc == True:
        return call(args, stdout=PIPE, stderr=STDOUT, creationflags=8) == 0
    else:
        return False

####################################################

#This function adjusts an IP by a supplied amount.
#Example AdjIp("10.10.10.0",1) = "10.10.10.1"
def adj_Ip(IP,adjby=0):
    """ """
    splitIP = IP.split(".")
    if len(splitIP) == 1:
        splitIP=IP.split(":")
    splitIP[-1] = int(splitIP[-1])
    splitIP[-1] += adjby
    splitIP[-1] = str(splitIP[-1])
    if len(splitIP) == 4:
        newIP = ".".join(splitIP)
    else:
        newIP = ":".join(splitIP)
    return newIP

####################################################

#This function takes an IP list and determines which IPs are alive
#and which are dead.
def DeadIPFilter(gatewayIPlist):
    aliveIPlist = []
    deadSubnetlist = []
    for IP in gatewayIPlist:
        if is_network_alive(IP,dc=True):
            aliveIPlist.append(IP)
        else:
            deadSubnetlist.append(adj_Ip(IP,-1))
            
    return aliveIPlist,deadSubnetlist            
            
#################################################    

#This function takes in an IP address and either returns a hostname or
#notes there is no DNS registration with a return of False.
def run_DNS(IP):
    try:
        name, alias, addresslist = socket.gethostbyaddr(IP)
    except:
        return False  
    else:
        return name

#################################################  

#This function takes in a list of IP addresses and tests for DNS registration.
#Results are sorted and returned accordingly.
def list_DNS(IPlist):
    HostnameList = []
    noDNSlist = []
    
    for IP in IPlist:
        hostname = run_DNS(IP)
        if hostname:
            HostnameList.append(hostname)
        else:
            noDNSlist.append(IP)
                     
    HostnameList.sort()
    noDNSlist.sort()
    
    return HostnameList,noDNSlist
    
##################################################

#This function extracts a device name from a full DNS hostname of an interface IP.
def HN_extract(hostname,term,adj_index_by=0):
    try:
        NameEndIndex = hostname.index(term) + adj_index_by
        extract_name = hostname[:NameEndIndex]
    except:
        return "ERROR"
    else:
        return extract_name

##################################################

#This function logs into a device and grabs the local device name.
#It can be used to test SSH or to determine a device name for an IP
#without DNS registration.  This process is much slower than DNS queries so it
#is used as an alternative method or after duplicate device names are removed.  

#It is also leveraged by other higher level functions.
def get_device_name(item):
    NetDevice = NetDev(item)
    NetDevice.login(quiet=True) 
    output = NetDevice.sendcmd("\n")[0].strip()[:-1]
    NetDevice.logout(quiet=True)
    return output+".cisco.com"
    
##################################################

#Takes a list of device names and turns them into a list of network objects.
#It also runs a few methods which create in-scope interface lists for each OOP instance.
def make_net_obj_list(DeviceList):
    NetObjList = [NetDev(device) for device in DeviceList]
    return NetObjList
	
##################################################

#This is the main network object class.  More granular device type child classes
#can inherit from this common parent class.       
class NetDev:
    """ Generic Network Device """

    def __init__(self,name):
        self.name = name
        self.loggedin = False

    def login(self,quiet=False): 
        if not self.loggedin:
            username,password = get_login()
            if username == None and password == None:
                set_login()
                username,password = get_login()
        
            # Create instance of SSHClient object
            self.ssh = paramiko.SSHClient()
        
            # Automatically add untrusted hosts (make sure okay for security policy in your environment)
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
            # initiate SSH connection
            self.ssh.connect(self.name, username=username, password=password, look_for_keys=False, allow_agent=False)
        
            # Use invoke_shell to establish an 'interactive session'
            self.remote_conn = self.ssh.invoke_shell()
            time.sleep(2)
               
            self.sendcmd("terminal length 0\n",buff=10000)
        
            if not quiet:
                print "Login to %s successful." % self.name
            self.loggedin = True
        else:
            print "You are already logged into %s." % self.name
        
    def logout(self,quiet=False):
        if self.loggedin:
            self.ssh.close()
            self.loggedin = False
            if not quiet:
                print "Successfully logged out of %s." % self.name
        else:
            print "You are not logged into %s." % self.name
        
    def sendcmd(self,command,sleep=1,buff=5000):
        # Now let's try to send the router a command
        self.remote_conn.send(command+"\n")
        # Wait for the command to complete
        time.sleep(sleep)
        self.output = self.remote_conn.recv(buff).split("\r\n")

        while True:
            try:
                index = self.output.index('')
                del self.output[index]
            except:
                break
        
        return self.output

    def TcpPortOpen(self,TcpPort):
        try:
            tn = telnetlib.Telnet(self.name,str(TcpPort))
            tn.close()
        except:
            return False
        else:
            return True
			
##################################################

#This child class represents a router with some device type specific methods.
class Router(NetDev):
    """Router Object"""
    pass
                
##################################################

class IPv4:
    def __init__(self,address):
        self.address = address    

##################################################

#This function creates an XLSX file from a network device object list.
def make_cfg_xlsx(filename,NetObjList):
    wb = openpyxl.Workbook()
    for i,NetDevice in enumerate(NetObjList):
        wb.create_sheet(index=i, title=NetDevice.name)
        sheet = wb.get_sheet_by_name(NetDevice.name)
        for i,line in enumerate(NetDevice.config):
            cell = "A"+str(i+1)
            sheet[cell] = line
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(filename)
            
##################################################

def read_xlsx(xlfile,xltab,xlcolumn="A",xlrow=2):
    filelist = []
    
    wb = openpyxl.load_workbook(xlfile)
    sheet = wb.get_sheet_by_name(xltab)
    xlvalue = True
    while xlvalue != "None":
        xlcell = xlcolumn+str(xlrow)
        xlvalue = str(sheet[xlcell].value).strip()
        filelist.append(xlvalue)
        xlrow +=1
    
    return filelist[:-1]

##################################################    

#This function reads a text file and returns a list.
def read_txt(filename):
    file = open(filename)
    FileList = file.readlines()
    FileList = [x.strip() for x in FileList]
    file.close()
    return FileList
	
##################################################
    
#This function takes a list and creates a text file.
def write_txt(filename,mylist):
    myfile = open(filename,"w")
    for item in mylist:
        myfile.write(item+"\n")
    myfile.close()

##################################################

def set_login():
    username = getuser()
    username = raw_input("Username '%s':" % username)
    if not username:
        username = getuser()
    password = getpass("CEC Password: ")
    f = open("c:\\pylogin.dat","wb")
    pickle.dump(username,f)
    pickle.dump(password,f)
    f.close()

##################################################

def get_login():
    try:
        f = open("c:\\pylogin.dat","rb")
        username = pickle.load(f)
        password = pickle.load(f)
        f.close()
    except:
        print "Login Credentials not set."
        return None,None
    else:
        return username,password

##################################################

def get_EMAN_dev_list(pgm=None,service=None):
    if pgm == None:
        pgm = raw_input("Enter PGM: ").lower()
    if service == None:
        service = raw_input("Enter Service: ").lower()
    p = Popen("eman-cli host find * --pgm=%s --service=%s" % (pgm,service), stdout=PIPE,stderr=STDOUT, shell=True)
    (output, err) = p.communicate()
    return output.split("\r\n")[:-2]
    
##################################################

def in_pre_not_post(file1,file2):
    list1 = read_txt(file1)
    list2 = read_txt(file2)
    
    diff_list = [item for item in list1 if item not in list2]
    return diff_list
    
##################################################

